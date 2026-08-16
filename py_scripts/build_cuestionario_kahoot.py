#!/usr/bin/env python3
"""
Genera un .xlsx en el formato exacto de importación de Kahoot a partir de una
lista de preguntas. Punto de partida: copia este archivo, edita QUESTIONS y
OUTPUT_PATH, y ejecuta con `python3 build_cuestionario_kahoot.py`.

Reglas de Kahoot (no las cambies sin verificar contra
https://support.kahoot.com primero):
  - Pregunta: máx. 95 caracteres.
  - Cada respuesta: máx. 60 caracteres.
  - Mínimo 2 respuestas por pregunta (Verdadero/Falso usa solo 2).
  - Tiempo permitido (seg): 5, 10, 20, 30, 60 o 120 (cualquier otro valor cae
    a 20 seg automáticamente al importar).
  - "Correct answer(s)": número(s) de respuesta correcta separados por coma,
    1-indexado (ej. "2" o "1,3").
"""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# ---------- EDITA ESTO PARA CADA GUÍA ----------
TITLE_SHEET_INFO = "CUESTIONARIO - CLEI III - Guía 1 / Semana 2"
OUTPUT_PATH = "/sessions/determined-sharp-johnson/mnt/outputs/Cuestionario_Kahoot.xlsx"

# Cada tupla: (pregunta, resp1, resp2, resp3_o_"", resp4_o_"", tiempo_seg, correcta(s) "1" o "2,3")
# Reglas de contenido: 7 preguntas de selección múltiple (4 opciones) + 3 de
# Verdadero/Falso (2 opciones, dejar resp3/resp4 en ""), todas basadas en los
# conceptos exactos que trae la guía de esa semana — no inventes temas nuevos.
QUESTIONS = [
    # (pregunta, r1, r2, r3, r4, tiempo, correcta)
]
# ------------------------------------------------

def build(questions, output_path, sheet_info_title):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Import Questions"

    headers = ["Question", "Answer 1", "Answer 2", "Answer 3", "Answer 4", "Time limit (sec)", "Correct answer(s)"]
    ws.append(headers)
    for r in questions:
        ws.append(list(r))

    # Validación de límites de Kahoot antes de guardar
    for i, r in enumerate(questions, start=2):
        q = r[0]
        assert len(q) <= 95, f"Fila {i}: pregunta de {len(q)} caracteres (máx 95): {q}"
        for a in r[1:5]:
            assert len(str(a)) <= 60, f"Fila {i}: respuesta demasiado larga: {a}"
        assert r[5] in (5, 10, 20, 30, 60, 120), f"Fila {i}: tiempo {r[5]} no permitido por Kahoot"

    font = Font(name="Arial", size=11)
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1B5B2A", end_color="1B5B2A", fill_type="solid")

    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(headers)):
        for cell in row:
            cell.font = font
            cell.alignment = Alignment(vertical="center", wrap_text=True, horizontal="left" if cell.column == 1 else "center")

    widths = [55, 18, 18, 18, 18, 14, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 30

    ws2 = wb.create_sheet("Instrucciones")
    legend = [
        (sheet_info_title, True),
        ("", False),
        ("Cómo importar este archivo en Kahoot:", True),
        ("1. Ingresa a create.kahoot.it y haz clic en Crear > Kahoot > Lienzos en blanco.", False),
        ('2. Haz clic en "Añadir" > pestaña "Importar" > "Importar hoja de cálculo".', False),
        ('3. Sube este archivo .xlsx (hoja "Import Questions") y confirma con "Cargar".', False),
        ('4. Clic en "Añadir preguntas". Elimina la pregunta 1 en blanco que Kahoot deja por defecto.', False),
        ("5. Ponle título al kahoot, marca visibilidad Privado y Guarda.", False),
        ("", False),
        ("Reglas del formato (Kahoot):", True),
        ("- Máximo 95 caracteres por pregunta.", False),
        ("- Máximo 60 caracteres por respuesta.", False),
        ("- Cada pregunta necesita mínimo 2 respuestas.", False),
        ("- Tiempo permitido (seg): 5, 10, 20, 30, 60 o 120.", False),
        ('- "Correct answer(s)" indica el número de la(s) respuesta(s) correcta(s), separadas por coma.', False),
    ]
    for i, (text, bold) in enumerate(legend, start=1):
        c = ws2.cell(row=i, column=1, value=text)
        c.font = Font(name="Arial", size=11, bold=bold)
    ws2.column_dimensions["A"].width = 100

    wb.save(output_path)
    print(f"Guardado: {output_path} ({len(questions)} preguntas)")


if __name__ == "__main__":
    if not QUESTIONS:
        raise SystemExit("Completa la lista QUESTIONS antes de ejecutar este script.")
    build(QUESTIONS, OUTPUT_PATH, TITLE_SHEET_INFO)
